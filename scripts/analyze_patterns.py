# -*- coding: utf-8 -*-
"""낙찰 패턴 다각도 분석 — bid_results → 세그먼트별 적정 투찰률 리포트 (xlsx).

세그먼트 축: 공종×하한율밴드 / 기관유형×공종 / 발주기관 / 금액대×공종 /
            다이제스트 관심그룹 / 기관별 사정율 편향 (용역 승부처)

지표: n, 하한율(대표), 사정율 중앙·IQR, 마진(낙찰률−하한율) p25/중앙/p75,
      참가업체수 중앙, 제안 투찰률 = 하한율 + 마진p25

Usage:
    python -m scripts.analyze_patterns [--out 경로.xlsx]
    (DATABASE_URL 또는 SUPABASE_URL+SUPABASE_SERVICE_ROLE_KEY — .env 자동 로드)
"""
from __future__ import annotations

import argparse
import os
import statistics
import sys
from collections import defaultdict
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(ROOT, ".env"))

from scripts.digest_v2 import GROUPS, _matches  # noqa: E402  (관심그룹 정의 단일화)

EOK = 100_000_000

# ── 기관유형 분류 (org_name 규칙 기반, 위에서부터 첫 매칭) ──
ORG_TYPES = [
    ("교육",     ["교육청", "교육지원청", "학교", "대학", "유치원"]),
    ("의료",     ["병원", "의료원", "보건"]),
    ("군·국방",  ["국방", "방위", "육군", "해군", "공군", "부대", "사단", "군단", "재정관리단"]),
    ("공기업·공단", ["공사", "공단", "공공기관", "진흥원", "연구원", "공항", "철도", "수자원", "토지주택", "전력"]),
    ("중앙부처", ["조달청", "국토관리", "지방청", "부 ", "청 ", "관리소", "국립"]),
    ("지자체",   ["특별시", "광역시", "특별자치", "도 ", "시 ", "군 ", "구 ", "시청", "군청", "구청", "면 ", "읍 "]),
    ("금융·조합", ["은행", "농협", "수협", "신협", "조합"]),
]


def org_type(name: str | None) -> str:
    n = (name or "") + " "
    for label, kws in ORG_TYPES:
        if any(k in n for k in kws):
            return label
    return "민간·기타"


def amount_band(base_price) -> str:
    if not base_price:
        return "미상"
    e = base_price / EOK
    if e < 5:
        return "5억 미만"
    if e < 10:
        return "5~10억"
    if e < 30:
        return "10~30억"
    if e < 100:
        return "30~100억"
    return "100억 이상"


def our_group(title: str, org: str) -> str | None:
    for gname, kws, g_exc, g_org_exc in GROUPS:
        if not _matches(title, kws):
            continue
        if g_exc and _matches(title, g_exc):
            continue
        if g_org_exc and _matches(org or "", g_org_exc):
            continue
        return gname
    return None


def fetch_results() -> list[dict]:
    cols = ("bid_no,title,org_name,bsns_div,decision_method,base_price,planned_price,"
            "sajeong_rate,win_lower_rate,win_bid_rate,bidder_count,open_result_date")
    url = os.environ.get("DATABASE_URL", "")
    if url:
        import psycopg2
        from psycopg2.extras import RealDictCursor
        if "sslmode=" not in url:
            url = f"{url}{'&' if '?' in url else '?'}sslmode=require"
        conn = psycopg2.connect(url, cursor_factory=RealDictCursor)
        try:
            with conn.cursor() as cur:
                cur.execute(f"SELECT {cols} FROM bid_results")
                return [dict(r) for r in cur.fetchall()]
        finally:
            conn.close()
    import requests
    base = os.environ["SUPABASE_URL"].rstrip("/")
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    hdr = {"apikey": key, "Authorization": f"Bearer {key}"}
    rows, offset = [], 0
    while True:
        r = requests.get(f"{base}/rest/v1/bid_results",
                         headers={**hdr, "Range-Unit": "items", "Range": f"{offset}-{offset + 999}"},
                         params={"select": cols, "order": "id.asc"}, timeout=30)
        r.raise_for_status()
        chunk = r.json()
        rows.extend(chunk)
        if len(chunk) < 1000:
            return rows
        offset += 1000


def _q(vals, p):
    vals = sorted(vals)
    if not vals:
        return None
    i = (len(vals) - 1) * p
    lo, hi = int(i), min(int(i) + 1, len(vals) - 1)
    return vals[lo] + (vals[hi] - vals[lo]) * (i - lo)


def build_population(rows: list[dict]) -> list[dict]:
    pop = []
    for x in rows:
        if "적격" not in (x.get("decision_method") or ""):
            continue
        lw, wr = x.get("win_lower_rate"), x.get("win_bid_rate")
        if not lw or not wr:
            continue
        lw, wr = float(lw), float(wr)
        m = wr - lw
        if not (0 <= m <= 15):
            continue
        sj = float(x["sajeong_rate"]) if x.get("sajeong_rate") else None
        if sj is not None and not (95 <= sj <= 105):
            continue
        pop.append({
            **x, "_lw": lw, "_wr": wr, "_m": m, "_sj": sj,
            "_otype": org_type(x.get("org_name")),
            "_band": amount_band(x.get("base_price")),
            "_group": our_group(x.get("title") or "", x.get("org_name") or ""),
        })
    return pop


def seg_stats(items: list[dict]) -> dict:
    ms = [x["_m"] for x in items]
    sj = [x["_sj"] for x in items if x["_sj"] is not None]
    bc = [x["bidder_count"] for x in items if x.get("bidder_count")]
    lw = [x["_lw"] for x in items]
    lower = statistics.median(lw)
    p25 = _q(ms, 0.25)
    return {
        "n": len(items),
        "하한율": round(lower, 3),
        "사정율중앙": round(statistics.median(sj), 3) if sj else None,
        "사정율IQR": round(_q(sj, 0.75) - _q(sj, 0.25), 3) if len(sj) >= 4 else None,
        "마진p25": round(p25, 3),
        "마진중앙": round(statistics.median(ms), 3),
        "마진p75": round(_q(ms, 0.75), 3),
        "참가중앙": statistics.median(bc) if bc else None,
        "제안투찰률": round(lower + p25, 3),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="", help="xlsx 출력 경로 (기본: data/낙찰패턴_리포트_날짜.xlsx)")
    ap.add_argument("--min-n", type=int, default=5)
    args = ap.parse_args()

    rows = fetch_results()
    pop = build_population(rows)
    print(f"bid_results {len(rows)}공고 → 적격심사 유효 {len(pop)}건, "
          f"관심그룹 매칭 {sum(1 for x in pop if x['_group'])}건")

    sheets: dict[str, list] = {}

    def emit(sheet: str, keyfn, min_n: int, top: int | None = None):
        groups = defaultdict(list)
        for x in pop:
            k = keyfn(x)
            if k:
                groups[k].append(x)
        out = [(k, seg_stats(v)) for k, v in groups.items() if len(v) >= min_n]
        out.sort(key=lambda kv: -kv[1]["n"])
        if top:
            out = out[:top]
        sheets[sheet] = out
        return out

    emit("공종x하한율", lambda x: f'{x["bsns_div"]} {round(x["_lw"], 3)}', args.min_n)
    emit("기관유형x공종", lambda x: f'{x["_otype"]} · {x["bsns_div"]}', args.min_n)
    emit("금액대x공종", lambda x: f'{x["_band"]} · {x["bsns_div"]}', args.min_n)
    emit("발주기관", lambda x: x.get("org_name"), args.min_n, top=60)
    emit("관심그룹", lambda x: x["_group"], 2)
    emit("관심그룹x기관유형", lambda x: f'{x["_group"]} · {x["_otype"]}' if x["_group"] else None, 3)

    # 기관별 사정율 편향 (용역 승부처 — 예측 정밀도 소재)
    sj_groups = defaultdict(list)
    for x in pop:
        if x["_sj"] is not None:
            sj_groups[x.get("org_name")].append(x["_sj"])
    sj_out = [(k, {"n": len(v), "사정율중앙": round(statistics.median(v), 3),
                   "IQR": round(_q(v, 0.75) - _q(v, 0.25), 3),
                   "편향(100대비)": round(statistics.median(v) - 100, 3)})
              for k, v in sj_groups.items() if len(v) >= args.min_n]
    sj_out.sort(key=lambda kv: abs(kv[1]["편향(100대비)"]), reverse=True)
    sheets["기관별사정율편향"] = sj_out[:60]

    # ── xlsx 출력 ──
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    out_path = args.out or os.path.join(ROOT, "data", f"낙찰패턴_리포트_{date.today():%Y%m%d}.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for sheet, data in sheets.items():
        ws = wb.create_sheet(sheet[:28])
        if not data:
            continue
        headers = ["구분"] + list(data[0][1].keys())
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2A78D6")
        for k, s in data:
            ws.append([str(k)] + list(s.values()))
        ws.column_dimensions["A"].width = 42
        for i in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 11
        ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"리포트 저장: {out_path}")

    print("\n── 관심그룹 요약 ──")
    for k, s in sheets["관심그룹"]:
        print(f"  {k:<12} n={s['n']:>4} 하한 {s['하한율']} 마진 {s['마진p25']}/{s['마진중앙']}/{s['마진p75']} → 제안 {s['제안투찰률']}%")
    return 0


if __name__ == "__main__":
    sys.exit(main())
