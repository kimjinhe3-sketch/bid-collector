"""Test dashboard helper functions without launching Streamlit runtime."""
from datetime import date, timedelta

import pandas as pd

from dashboard import app as dashboard
from db import database


# ─────────────── Region extraction (④) ───────────────

def test_extract_region_basic():
    assert dashboard._extract_region("서울특별시 강남구") == "서울"
    assert dashboard._extract_region("경기도 화성시") == "경기"
    assert dashboard._extract_region("부산광역시 해운대구") == "부산"
    assert dashboard._extract_region("충청북도 청주시") == "충북"
    assert dashboard._extract_region("전북특별자치도 전주시") == "전북"
    assert dashboard._extract_region("강원특별자치도 춘천시") == "강원"


def test_extract_region_unknown_returns_default():
    assert dashboard._extract_region("한국수자원공사") == "전국/기타"
    assert dashboard._extract_region("") == "전국/기타"
    assert dashboard._extract_region(None) == "전국/기타"


# ─────────────── D-n countdown (③) ───────────────

def test_dn_label_categories():
    today = date(2026, 5, 1)
    # 마감 지남
    assert dashboard._dn_label("2026-04-30", today) == "마감"
    # 오늘 마감
    assert dashboard._dn_label("2026-05-01", today) == "D-day"
    # D-1 ~ D-7
    assert dashboard._dn_label("2026-05-02", today) == "D-1"
    assert dashboard._dn_label("2026-05-03 14:00", today) == "D-2"
    assert dashboard._dn_label("2026-05-08", today) == "D-7"
    # 8일 이상 → 빈 문자열
    assert dashboard._dn_label("2026-05-09", today) == ""
    # 파싱 실패
    assert dashboard._dn_label("garbage", today) == ""
    assert dashboard._dn_label(None, today) == ""


def test_days_until():
    today = date(2026, 5, 1)
    assert dashboard._days_until("2026-05-04", today) == 3
    assert dashboard._days_until("2026-04-28", today) == -3
    assert dashboard._days_until("2026-05-01 23:59", today) == 0
    assert dashboard._days_until("invalid", today) is None


def test_rows_to_dataframe_includes_region_and_dday_columns():
    rows = [_row(1, source="g2b_api_thng")]
    rows[0]["org_name"] = "경기도 성남시"
    rows[0]["close_date"] = (date.today() + timedelta(days=2)).strftime("%Y-%m-%d")
    df = dashboard.rows_to_dataframe(rows)
    assert "지역" in df.columns
    assert "마감임박" in df.columns
    assert df["지역"].iloc[0] == "경기"
    assert df["마감임박"].iloc[0] == "D-2"


def _row(n, source="g2b_api_thng", price=500_000_000, bid_type="용역"):
    return {
        "source": source,
        "bid_no": f"B{n}",
        "title": f"sample {n}",
        "org_name": "기관",
        "contract_method": "제한경쟁",
        "estimated_price": price,
        "open_date": "2026-04-22 10:00",
        "close_date": "2026-05-02 10:00",
        "bid_type": bid_type,
        "detail_url": f"https://example.com/{n}",
    }


def test_rows_to_dataframe_empty():
    df = dashboard.rows_to_dataframe([])
    assert isinstance(df, pd.DataFrame)
    assert len(df) == 0
    assert "금액(억원)" in df.columns
    assert "title" in df.columns
    assert "신규" in df.columns


def test_rows_to_dataframe_new_badge_today_only():
    from datetime import date, timedelta
    today = date.today()
    today_str = today.strftime("%Y-%m-%d")
    yesterday_str = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    rows = [
        {**_row(1), "open_date": today_str + " 10:00"},
        {**_row(2), "open_date": yesterday_str},
        {**_row(3), "open_date": None},
    ]
    df = dashboard.rows_to_dataframe(rows)
    vals = list(df["신규"])
    assert vals[0] == "new"   # 오늘 → new
    assert vals[1] == ""      # 어제 → 공란
    assert vals[2] == ""      # 파싱 실패 → 공란


def test_rows_to_dataframe_formats_amount_column():
    import math
    rows = [
        {**_row(1), "estimated_price": 850_000_000},
        {**_row(2), "estimated_price": 24_150_000},
        {**_row(3), "estimated_price": None},
        {**_row(4), "estimated_price": 0},
    ]
    df = dashboard.rows_to_dataframe(rows)
    vals = list(df["금액(억원)"])
    # 숫자형 (억 단위), 2자리 소수
    assert vals[0] == 8.50
    assert vals[1] == 0.24
    # None/0 → NaN (정렬 시 pandas 기본으로 최하단)
    assert vals[2] is None or (isinstance(vals[2], float) and math.isnan(vals[2]))
    assert vals[3] is None or (isinstance(vals[3], float) and math.isnan(vals[3]))


def test_rows_to_dataframe_maps_source_label():
    rows = [_row(1, source="g2b_api_servc"), _row(2, source="alio")]
    df = dashboard.rows_to_dataframe(rows)
    assert "나라장터 용역" in list(df["source_label"])
    assert "ALIO" in list(df["source_label"])


def test_rows_to_dataframe_upgrades_kepco_http_to_https():
    rows = [
        {**_row(1, source="kepco_api"), "title": "KEPCO 공고",
         "detail_url": "http://srm.kepco.net/printDownloadAttachment.do?id=abc"},
        {**_row(2, source="kepco_api"), "title": "KEPCO 공고 2",
         "detail_url": "https://srm.kepco.net/printDownloadAttachment.do?id=xyz"},
    ]
    df = dashboard.rows_to_dataframe(rows)
    urls = list(df["detail_url"])
    assert urls[0] == "https://srm.kepco.net/printDownloadAttachment.do?id=abc"
    assert urls[1] == "https://srm.kepco.net/printDownloadAttachment.do?id=xyz"


def test_df_to_excel_bytes_returns_valid_xlsx():
    """df_to_excel_bytes 는 openpyxl 로 읽을 수 있는 유효한 xlsx 바이트를 반환.
    헤더/하이퍼링크/NEW 배지 포맷이 포함되어야 함.
    """
    import io
    from openpyxl import load_workbook
    from datetime import date, timedelta

    today = date.today()
    rows = [
        {**_row(1), "title": "오늘 공고",
         "open_date": today.strftime("%Y-%m-%d %H:%M"),
         "estimated_price": 850_000_000,
         "detail_url": "https://example.com/1"},
        {**_row(2), "title": "어제 공고",
         "open_date": (today - timedelta(days=1)).strftime("%Y-%m-%d"),
         "estimated_price": None,
         "detail_url": None},
    ]
    df = dashboard.rows_to_dataframe(rows)
    buf = dashboard.df_to_excel_bytes(df)
    assert isinstance(buf, bytes) and len(buf) > 1000

    wb = load_workbook(io.BytesIO(buf))
    ws = wb.active
    # 헤더 첫 행
    headers = [c.value for c in ws[1]]
    assert "공고번호" in headers
    assert "제목" in headers
    assert "금액(억원)" in headers
    assert "상세보기(링크)" in headers
    # 2행(오늘 공고): NEW 배지 셀 = 'new' 값, hyperlink 존재
    assert ws.cell(row=2, column=headers.index("신규") + 1).value == "new"
    link_cell = ws.cell(row=2, column=headers.index("상세보기(링크)") + 1)
    assert link_cell.hyperlink is not None
    assert link_cell.hyperlink.target == "https://example.com/1"
    assert link_cell.value == "열기"
    # 3행(어제): 신규 셀 공백, detail_url None → 일반 셀
    assert ws.cell(row=3, column=headers.index("신규") + 1).value in ("", None)
    assert ws.freeze_panes == "A2"


def test_rows_to_dataframe_reconstructs_prvt_pdf_url_when_missing():
    """Legacy 누리장터 rows may have detail_url=None. rows_to_dataframe
    should reconstruct the 공고문 PDF download URL from bid_no so the link
    column is always clickable. URL format mirrors what the API returns.
    """
    rows = [
        {**_row(1, source="prvt_api_servc"), "title": "ABC 교회 태양광",
         "bid_no": "R26BK01482245-000", "detail_url": None},
        {**_row(2, source="prvt_api_thng"), "title": "XX 아파트 승강기",
         "bid_no": "R26BK01500000", "detail_url": ""},
        # Existing non-empty url should NOT be replaced
        {**_row(3, source="prvt_api_cnstwk"), "title": "YY 공사",
         "bid_no": "R26X", "detail_url": "https://example.com/real"},
        # Non-prvt row with empty url stays empty
        {**_row(4, source="g2b_api_thng"), "detail_url": None},
    ]
    df = dashboard.rows_to_dataframe(rows)
    urls = list(df["detail_url"])
    # Legacy prvt with -ord suffix → properly split
    assert "bidPbancNo=R26BK01482245" in urls[0]
    assert "bidPbancOrd=000" in urls[0]
    assert "prcmBsneSeCd=22" in urls[0]
    # No -ord suffix → default 000 used
    assert "bidPbancNo=R26BK01500000" in urls[1]
    assert "bidPbancOrd=000" in urls[1]
    assert urls[0].startswith("https://www.g2b.go.kr/pn/pnp/pnpe/UntyAtchFile/")
    # Existing URL not overridden
    assert urls[2] == "https://example.com/real"
    # g2b row with None stays None/NaN — not touched
    import math
    u3 = urls[3]
    is_empty = (u3 is None or u3 == ""
                or (isinstance(u3, float) and math.isnan(u3)))
    assert is_empty or (isinstance(u3, str) and "g2b.go.kr/pn/pnp" not in u3)


def test_rows_to_dataframe_rewrites_stale_alio_bidview_urls():
    rows = [
        {
            **_row(1, source="alio"),
            "title": "AI 기반 데이터 플랫폼",
            "detail_url": "https://www.alio.go.kr/occasional/bidView.do?seq=3520838",
        },
        {
            **_row(2, source="alio"),
            "title": "이미 정상",
            "detail_url": "https://www.alio.go.kr/occasional/bidList.do?type=title&word=ok",
        },
        {
            **_row(3, source="g2b_api_thng"),
            "title": "G2B 건드리지 말 것",
            "detail_url": "https://www.g2b.go.kr/foo?seq=1",
        },
    ]
    df = dashboard.rows_to_dataframe(rows)
    urls = list(df["detail_url"])
    # 레거시 ALIO URL은 search URL로 변환됨
    assert "bidView.do" not in urls[0]
    assert "bidList.do?type=title" in urls[0]
    # 이미 정상인 ALIO URL은 그대로
    assert urls[1] == "https://www.alio.go.kr/occasional/bidList.do?type=title&word=ok"
    # G2B URL은 건드리지 않음
    assert urls[2] == "https://www.g2b.go.kr/foo?seq=1"


def test_rows_to_dataframe_uses_raw_source_when_unknown():
    rows = [_row(1, source="unknown_source")]
    df = dashboard.rows_to_dataframe(rows)
    assert list(df["source_label"]) == ["unknown_source"]


def test_source_labels_cover_known_sources():
    expected = {
        "g2b_api_thng", "g2b_api_servc", "g2b_api_cnstwk",
        "g2b_api_frgcpt", "g2b_api_etc",
        "prvt_api_servc", "prvt_api_thng",
        "prvt_api_cnstwk", "prvt_api_etc",
        "alio", "g2b_crawl",
        "d2b_api_dmstc", "kwater_api",
        "kwater_api_cntrwk", "kwater_api_gds",
        "kwater_api_servc", "kwater_api_dmscpt",
        "kepco_api",
        "lh_api", "kec_api",   # 신규: 한국토지주택공사 + 한국도로공사
    }
    assert expected == set(dashboard.SOURCE_LABELS.keys())


def test_load_rows_uses_database_helpers(tmp_path, monkeypatch):
    """Ensure the cached loader calls database.fetch_for_dashboard with right args."""
    db = tmp_path / "t.sqlite"
    database.init_db(db)
    database.upsert_bids(db, [
        _row(1, bid_type="용역"),
        _row(2, bid_type="물품"),
        _row(3, bid_type="공사"),
    ])
    dashboard.load_rows.clear()  # clear st.cache_data between tests
    out = dashboard.load_rows(str(db), None, ("용역", "물품"), None, 100)
    assert {r["bid_type"] for r in out} == {"용역", "물품"}

    dashboard.load_rows.clear()
    out = dashboard.load_rows(str(db), None, (), "sample 2", 100)
    assert len(out) == 1
    assert out[0]["bid_no"] == "B2"
