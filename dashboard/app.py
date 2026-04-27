"""대한민국 입찰공고 대시보드 (Streamlit).

로컬:  streamlit run dashboard/app.py
배포:  Streamlit Community Cloud — 이 파일을 entrypoint로 지정.

환경변수/시크릿은 utils.secrets.get_secret()로 통일 해석.
"""
from __future__ import annotations

import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

# Windows corporate proxy SSL fix — no-op on Linux (Streamlit Cloud).
try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

import pandas as pd
import streamlit as st

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

# Load .env when running locally (no-op on Streamlit Cloud).
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

from db import database
from filters import keyword_filter
from utils.config_loader import load_config
from utils.secrets import get_secret
from utils import recipients as recipients_mod


EOK = 100_000_000

# 메트릭 카드용 상위 그룹: 3개 버킷으로 집계
SOURCE_GROUPS = {
    "나라장터": [
        "g2b_api_thng", "g2b_api_servc", "g2b_api_cnstwk",
        "g2b_api_frgcpt", "g2b_api_etc",
    ],
    "누리장터": [
        "prvt_api_servc", "prvt_api_thng",
        "prvt_api_cnstwk", "prvt_api_etc",
    ],
    "기타": [
        "d2b_api_dmstc", "kepco_api", "alio",
        "kwater_api", "kwater_api_cntrwk", "kwater_api_gds",
        "kwater_api_servc", "kwater_api_dmscpt", "g2b_crawl",
    ],
}

SOURCE_LABELS = {
    "g2b_api_thng": "나라장터 물품",
    "g2b_api_servc": "나라장터 용역",
    "g2b_api_cnstwk": "나라장터 공사",
    "g2b_api_frgcpt": "나라장터 외자",
    "g2b_api_etc":    "나라장터 기타",
    "prvt_api_servc":  "누리장터 용역",
    "prvt_api_thng":   "누리장터 물품",
    "prvt_api_cnstwk": "누리장터 공사",
    "prvt_api_etc":    "누리장터 기타",
    "alio": "ALIO",
    "g2b_crawl": "나라장터 크롤",
    "d2b_api_dmstc": "방위사업청",
    "kwater_api": "K-water",
    "kwater_api_cntrwk": "K-water 공사",
    "kwater_api_gds":    "K-water 물품",
    "kwater_api_servc":  "K-water 용역",
    "kwater_api_dmscpt": "K-water 내자",
    "kepco_api":         "KEPCO",
}


# ────────────────────────────────────────────────────────────────
# Styling — custom CSS incl. mobile media queries
# ────────────────────────────────────────────────────────────────

CUSTOM_CSS = """
<style>
/* ─────────────── Design spec (warm coral palette) ─────────────── */
:root {
  /* Primary action */
  --color-primary: #D85A30;
  --color-primary-hover: #993C1D;
  --color-border: #D3D1C7;
  /* Tags */
  --tag-include-bg: #FAECE7;
  --tag-include-text: #993C1D;
  --tag-include-border: #F5C4B3;
  --tag-exclude-bg: #F1EFE8;
  --tag-exclude-text: #5F5E5A;
  --tag-exclude-border: #D3D1C7;
  /* Typography hierarchy */
  --text-primary: #2C2C2A;    /* H1/H2 */
  --text-heading: #444441;    /* H3 */
  --text-body: #5F5E5A;       /* 본문 */
  --text-muted: #888780;      /* 보조 */
  /* Accent */
  --link: #185FA5;
  --accent-number: #D85A30;
  /* Surfaces */
  --bg-page: #FAF9F5;
  --bg-surface: #FFFFFF;
  --bg-sidebar: #F1EFE8;
  --border-line: #D3D1C7;
  /* Legacy aliases (기존 코드 호환) */
  --bg: #FAF9F5;
  --bg-soft: #F1EFE8;
  --card: #FFFFFF;
  --fg: #2C2C2A;
  --fg-muted: #888780;
  --border: #D3D1C7;
  --border-strong: #B8B5AC;
  --accent: #D85A30;
  --accent-hover: #993C1D;
  --accent-soft: #FAECE7;
  --danger: #993C1D;
  --success: #7C8F52;
  --radius-sm: 8px;
  --radius: 10px;
  --radius-lg: 14px;
}

html, body, .stApp { background: var(--bg) !important; }

/* 한글 친화 폰트 — 전역 기본만 설정 (Material Icons는 건드리지 않음) */
body, .stApp {
  font-family: "Pretendard Variable", Pretendard, "Apple SD Gothic Neo",
               Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  color: var(--fg);
  letter-spacing: -0.003em;
  -webkit-font-smoothing: antialiased;
}

/* 명시적 요소들만 !important — 아이콘 span은 제외 */
.stMarkdown, .stMarkdown p, .stMarkdown li, .stMarkdown span,
.stTextInput input, .stTextArea textarea,
.stDateInput input, .stNumberInput input,
.stButton > button, .stDownloadButton > button,
[data-testid="stCheckbox"] label, label, p {
  font-family: "Pretendard Variable", Pretendard, "Apple SD Gothic Neo",
               Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
  color: var(--fg);
}

/* Material Icons / Symbols — Streamlit Cloud 모바일에서 자체 폰트가
   안 들어오는 경우 Google Fonts CDN 폴백 (위에서 link 로 preload).
   ligature 가 raw 텍스트로 보이는 것 방지. */
.material-icons, .material-symbols-outlined, .material-symbols-rounded,
span[class*="material-icons"], span[class*="material-symbols"],
[data-testid="stIconMaterial"], [data-testid*="Icon"] > span {
  font-family: "Material Symbols Outlined", "Material Symbols Rounded",
               "Material Icons" !important;
  font-feature-settings: "liga" !important;
  font-variation-settings: "FILL" 0, "wght" 400, "GRAD" 0, "opsz" 24 !important;
  font-weight: normal !important;
  font-style: normal !important;
  line-height: 1 !important;
  letter-spacing: normal !important;
  text-transform: none !important;
  display: inline-block !important;
  white-space: nowrap !important;
  word-wrap: normal !important;
  direction: ltr !important;
  -webkit-font-smoothing: antialiased !important;
}

/* Typography hierarchy per design spec */
h1, h2, .brand-title {
  color: var(--text-primary) !important;
  font-weight: 600 !important;
  letter-spacing: -0.015em !important;
}
h3 {
  color: var(--text-heading) !important;
  font-size: 1.05rem !important;
  font-weight: 600 !important;
  margin: 1.2rem 0 0.5rem 0 !important;
  letter-spacing: -0.01em !important;
}
p, .stMarkdown, body { color: var(--text-body); }
.section-hint, [data-testid="stCaptionContainer"] { color: var(--text-muted) !important; }

/* ─── Brand bar ─── */
.brand-bar {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: var(--radius-lg);
  padding: 18px 22px;
  margin-bottom: 20px;
  box-shadow: 0 1px 2px rgba(61, 57, 41, 0.04);
}
.brand-title { font-size: 1.4rem; margin: 0; }
.brand-sub { font-size: 0.82rem; color: var(--fg-muted); margin-top: 4px; }

/* ─── Section hints ─── */
.section-hint {
  color: var(--fg-muted); font-size: 0.82rem;
  margin: -0.3rem 0 0.75rem 0;
}

/* ─── Tag chips (warm coral spec) ─── */
.kw-chip {
  display: inline-block; padding: 3px 10px; margin: 3px 4px 3px 0;
  background: var(--tag-include-bg);
  color: var(--tag-include-text);
  border: 1px solid var(--tag-include-border);
  border-radius: 999px;
  font-size: 0.78rem; font-weight: 500;
  line-height: 1.5;
}
.kw-chip.ex {
  background: var(--tag-exclude-bg);
  color: var(--tag-exclude-text);
  border: 1px solid var(--tag-exclude-border);
}

/* ─── Empty state ─── */
.empty-state {
  text-align: center; padding: 44px 20px;
  background: var(--card); border: 1px dashed var(--border-strong);
  border-radius: var(--radius); color: var(--fg-muted);
}

/* ─── Buttons (Claude warm) ─── */
.stButton > button {
  border-radius: var(--radius-sm) !important;
  border: 1px solid var(--border) !important;
  background: var(--card) !important;
  color: var(--fg) !important;
  font-weight: 500 !important;
  font-size: 0.88rem !important;
  padding: 0.4rem 0.95rem !important;
  transition: all 0.12s ease !important;
  box-shadow: none !important;
}
.stButton > button:hover {
  border-color: var(--border-strong) !important;
  background: #fdfcf7 !important;
}
.stButton > button[kind="primary"],
.stButton > button[kind="primary"] *,
.stButton > button[kind="primary"] p,
.stButton > button[kind="primary"] span,
.stButton > button[kind="primary"] div {
  background: var(--accent) !important;
  color: #ffffff !important;
  border: 1px solid var(--accent) !important;
  -webkit-text-fill-color: #ffffff !important;
}
.stButton > button[kind="primary"] svg { fill: #ffffff !important; }
.stButton > button[kind="primary"]:hover,
.stButton > button[kind="primary"]:hover * {
  background: var(--accent-hover) !important;
  border-color: var(--accent-hover) !important;
  color: #ffffff !important;
  -webkit-text-fill-color: #ffffff !important;
}

/* ─── Inputs ─── */
.stTextInput input, .stTextArea textarea,
.stDateInput input, .stNumberInput input,
.stSelectbox > div > div, .stMultiSelect > div > div {
  border-radius: var(--radius-sm) !important;
  border: 1px solid var(--border) !important;
  background: var(--card) !important;
  font-size: 0.88rem !important;
  color: var(--fg) !important;
}

/* ─── Multiselect 선택된 태그 (업종 등) — 코랄 팔레트 ─── */
.stMultiSelect [data-baseweb="tag"],
[data-baseweb="tag"] {
  background-color: var(--tag-include-bg) !important;
  color: var(--tag-include-text) !important;
  border: 1px solid var(--tag-include-border) !important;
  border-radius: 999px !important;
  font-weight: 500 !important;
}
.stMultiSelect [data-baseweb="tag"] span,
[data-baseweb="tag"] span {
  color: var(--tag-include-text) !important;
  -webkit-text-fill-color: var(--tag-include-text) !important;
}
.stMultiSelect [data-baseweb="tag"] svg,
[data-baseweb="tag"] svg {
  fill: var(--tag-include-text) !important;
  color: var(--tag-include-text) !important;
}
.stTextInput input:focus, .stTextArea textarea:focus,
.stDateInput input:focus, .stNumberInput input:focus {
  border-color: var(--accent) !important;
  box-shadow: 0 0 0 3px var(--accent-soft) !important;
  outline: none !important;
}

/* ─── Sidebar (warm soft) ─── */
/* 폭/배경만 커스터마이즈. transform/visibility/display 는 건드리지 않아
   Streamlit 기본 접기/펼치기 애니메이션이 정상 동작하게 둔다. */
[data-testid="stSidebar"] {
  background: var(--bg-soft);
  border-right: 1px solid var(--border);
  width: 21rem;
  min-width: 21rem;
}
/* 접기/펼치기 버튼 항상 보이게 (버전별 test-id 모두 커버) */
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapsedControl"],
[data-testid="stSidebarCollapseButton"],
[data-testid="stExpandSidebarButton"],
[data-testid="baseButton-headerNoPadding"],
[data-testid="stBaseButton-headerNoPadding"] {
  visibility: visible !important;
  display: inline-flex !important;
  opacity: 1 !important;
  pointer-events: auto !important;
  /* 아이콘 폰트가 늦게 로드되어도 button 이 collapse 되지 않도록 최소 크기 보장 */
  min-width: 32px !important;
  min-height: 32px !important;
  align-items: center !important;
  justify-content: center !important;
}
/* 햄버거 button 안 SVG/icon span 도 최소 크기 보장 */
[data-testid="stExpandSidebarButton"] > *,
[data-testid="stSidebarCollapseButton"] > *,
[data-testid="stSidebarCollapsedControl"] > *,
[data-testid="collapsedControl"] > * {
  min-width: 18px;
  min-height: 18px;
}
/* 태블릿 (≤ 900px): 사이드바 살짝 좁게 */
@media (max-width: 900px) {
  [data-testid="stSidebar"] {
    width: 17rem;
    min-width: 17rem;
  }
}
/* 모바일 (≤ 640px): 사이드바 폭 + 햄버거 버튼 강제 노출 */
@media (max-width: 640px) {
  /* 사이드바가 펼쳐졌을 때 화면 폭 대부분 차지 */
  [data-testid="stSidebar"][aria-expanded="true"] {
    width: 85vw !important;
    min-width: 0 !important;
    max-width: 85vw !important;
  }
  /* 햄버거 버튼은 Streamlit 기본 위치 (stHeader 안 좌상단) 그대로.
     우리는 강조만 — 테두리/배경/아이콘 색만 부여하고 크기·위치는 건드리지 않음
     (이전에 position:fixed 로 옮기려다 button 이 0×0 이 됨) */
  [data-testid="collapsedControl"],
  [data-testid="stSidebarCollapsedControl"],
  [data-testid="stExpandSidebarButton"] {
    background: var(--bg-surface) !important;
    border: 1px solid var(--accent) !important;
    border-radius: 8px !important;
    box-shadow: 0 2px 6px rgba(216, 90, 48, 0.18) !important;
  }
  /* 아이콘 색만 강조 (크기는 Streamlit 기본 유지) */
  [data-testid="stExpandSidebarButton"] svg,
  [data-testid="stExpandSidebarButton"] [data-testid="stIconMaterial"] {
    color: var(--accent) !important;
  }
  /* 모바일에서 stHeader 가 표시 영역 위로 가도록 z-index */
  [data-testid="stHeader"] { z-index: 999 !important; }
  /* 사이드바 내부 폰트·여백 축소 */
  [data-testid="stSidebar"] .block-container {
    padding: 0.6rem 0.6rem !important;
  }
  [data-testid="stSidebar"] h3 {
    font-size: 0.7rem !important;
    margin: 0.8rem 0 0.3rem 0 !important;
  }
  [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] .stCheckbox label,
  [data-testid="stSidebar"] input,
  [data-testid="stSidebar"] textarea,
  [data-testid="stSidebar"] button {
    font-size: 0.8rem !important;
  }
  [data-testid="stSidebar"] .stSlider { font-size: 0.75rem; }
}
[data-testid="stSidebar"] h3 {
  font-family: "Pretendard Variable", Pretendard, sans-serif !important;
  font-size: 0.78rem !important;
  color: var(--fg-muted) !important;
  text-transform: uppercase;
  letter-spacing: 0.08em;
  font-weight: 600 !important;
  margin: 1.2rem 0 0.4rem 0 !important;
}

/* ─── Checkboxes ─── */
[data-testid="stCheckbox"] label { font-size: 0.9rem; font-weight: 500; }

/* ─── 소스 필터 섹션 헤딩 ─── */
.src-section-title {
  font-family: "Pretendard Variable", Pretendard, sans-serif;
  font-size: 1.05rem;
  font-weight: 600;
  color: var(--fg);
  margin: 0 0 0.4rem 0;
  padding: 0;
}

/* ─── 모바일: 소스 필터 섹션 컴팩트화 ─── */
@media (max-width: 768px) {
  /* 섹션 헤딩 숨김 — 카운트 캡션이 충분히 정보 전달 */
  .st-key-src_filter_section .src-section-title {
    display: none !important;
  }
  /* 컬럼 row 의 모든 child container margin 0, gap 축소 */
  .st-key-src_filter_section [data-testid="stHorizontalBlock"] {
    gap: 0.3rem !important;
    flex-wrap: nowrap !important;
  }
  .st-key-src_filter_section [data-testid="stColumn"] {
    min-width: 0 !important;
    padding: 0 !important;
  }
  .st-key-src_filter_section [data-testid="stElementContainer"] {
    margin: 0 !important;
    padding: 0 !important;
  }
  /* 체크박스 — 라벨 작게, padding 축소, 한 줄에 3개 */
  .st-key-src_filter_section [data-testid="stCheckbox"] {
    padding: 2px 0 !important;
  }
  .st-key-src_filter_section [data-testid="stCheckbox"] label {
    font-size: 0.72rem !important;
    line-height: 1.2 !important;
  }
  .st-key-src_filter_section [data-testid="stCheckbox"] label > div {
    gap: 4px !important;
  }
  /* 버튼 (전체/해제) — 작고 짧게 */
  .st-key-src_filter_section [data-testid="stButton"] button {
    padding: 4px 6px !important;
    font-size: 0.72rem !important;
    min-height: 30px !important;
    height: 30px !important;
    line-height: 1 !important;
  }
}

/* ─── 표 타이틀바 (윈도우 창 스타일 · 표에 일체화) ─── */
.st-key-bidtable_titlebar {
  background: transparent !important;
  border: none !important;
  border-radius: 0 !important;
  padding: 6px 2px !important;
  margin-bottom: 0 !important;
  min-height: 42px;
}
.st-key-bidtable_titlebar .tbl-title {
  font-size: 0.92rem;
  color: var(--fg);
  line-height: 1.3;
}
/* 타이틀바 내부 모든 stElementContainer 의 여백 제거 */
.st-key-bidtable_titlebar [data-testid="stElementContainer"] {
  margin: 0 !important;
  padding: 0 !important;
}
/* 타이틀바 바로 뒤에 오는 dataframe 을 시각적으로 '연결' */
.st-key-bidtable_titlebar + [data-testid="stElementContainer"] {
  margin-top: 0 !important;
}
.st-key-bidtable_titlebar + [data-testid="stElementContainer"] [data-testid="stDataFrame"] {
  border-radius: 0 !important;   /* 표 전체 직사각형 */
  margin-top: 0 !important;
}

/* ─── 엑셀 다운로드 아이콘 (우상단·작게·옅은 색) ─── */
[data-testid="stDownloadButton"] button {
  background: transparent !important;
  color: rgba(120, 95, 80, 0.6) !important;
  border: 1px solid rgba(216, 90, 48, 0.2) !important;
  padding: 2px 10px !important;
  font-size: 0.72rem !important;
  font-weight: 500 !important;
  min-height: 28px !important;
  height: 28px !important;
  border-radius: 8px !important;
  box-shadow: none !important;
  transition: all 0.15s ease;
}
[data-testid="stDownloadButton"] button:hover {
  background: rgba(216, 90, 48, 0.06) !important;
  color: var(--accent) !important;
  border-color: rgba(216, 90, 48, 0.45) !important;
}

/* ─── Dataframe (spec colors) ─── */
[data-testid="stDataFrame"] {
  border-radius: var(--radius) !important;
  overflow: hidden;
  border: 1px solid var(--border-line);
  background: var(--bg-surface);
}
/* dataframe 의 touch-action 은 강제하지 않음 (Streamlit 기본값 사용).
   - 가로 컬럼 스크롤: glide-data-grid 자체 핸들러가 처리 (CSS 로 막으면 안 됨)
   - 세로 페이지 스크롤: main() 의 JS 핸들러가 sticky direction 으로
     window capture 단계에서 grid 핸들러를 stopImmediatePropagation. */

/* 모바일 viewport 채우기 — 표 아래 빈 공간 제거.
   calc(100vh - 280px) 의 280px = stHeader(60) + 로고카드(80) + 필터row(40)
   + 캡션(20) + 타이틀바(40) + 마진들 (~40). 화면 회전·기기별 차이는
   min/max-height 로 가드. */
@media (max-width: 768px) {
  [data-testid="stDataFrame"] {
    height: calc(100vh - 280px) !important;
    min-height: 420px !important;
    max-height: calc(100vh - 200px) !important;
  }
  [data-testid="stDataFrame"] > div,
  [data-testid="stDataFrame"] > div > div {
    height: 100% !important;
  }
}
/* "열기" 링크 — 파랑 유지로 클릭 가능 신호 */
[data-testid="stDataFrame"] a {
  text-decoration: none;
  color: var(--link) !important;
  font-weight: 500;
}
/* 금액(숫자) 강조 — 정수 컬럼 오른쪽 정렬 셀 */
[data-testid="stDataFrame"] [role="cell"][data-testid*="Number"],
[data-testid="stDataFrame"] [data-col-index] [style*="text-align: right"] {
  color: var(--accent-number) !important;
  font-weight: 600;
  font-variant-numeric: tabular-nums;
}

/* ─── Slider ─── */
.stSlider [role="slider"] { background: var(--accent) !important; }
[data-testid="stStatus"] { border-radius: var(--radius) !important; }

/* ─── Hide Streamlit chrome (사이드바 toggle은 유지) ─── */
footer { visibility: hidden; }
#MainMenu { visibility: hidden; }
/* stToolbar 자체는 숨기지 않음 — 그 안에 stExpandSidebarButton(햄버거)가
   있어서 숨기면 모바일에서 사이드바를 열 수 없음. 대신 안의 deploy/share
   같은 Streamlit 자체 actions 만 숨김. */
[data-testid="stToolbarActions"] { display: none !important; }
[data-testid="stDecoration"] { display: none; }
[data-testid="stAppDeployButton"] { display: none; }
/* header는 그대로 — 사이드바 toggle이 거기 있음 */

/* ─── Mobile (≤ 768px) ─── */
@media (max-width: 768px) {
  /* 전체 body 폰트 축소 */
  html, body, .stApp { font-size: 14px !important; }
  .brand-bar { padding: 12px 14px; }
  .brand-title { font-size: 1.05rem !important; }
  .brand-sub { font-size: 0.78rem !important; }
  h1 { font-size: 1.3rem !important; }
  h2 { font-size: 1.15rem !important; }
  h3 { font-size: 1rem !important; }
  p, span, div { font-size: 0.85rem; }
  [data-testid="stHorizontalBlock"] { flex-wrap: wrap; }
  .desktop-only { display: none; }
  [data-testid="stDataFrame"] { font-size: 0.78rem; }
  [data-testid="stDataFrame"] th,
  [data-testid="stDataFrame"] td { padding: 4px 6px !important; }
  .block-container {
    padding: 0.7rem 0.6rem !important;
    max-width: 100vw !important;
  }
  .stButton button { font-size: 0.85rem !important; padding: 0.4rem 0.8rem !important; }
  .kw-chip { font-size: 0.72rem !important; padding: 1px 6px !important; }
  [data-testid="stMetricValue"] { font-size: 1.1rem !important; }
  [data-testid="stMetricLabel"] { font-size: 0.75rem !important; }
}
/* ─── Small mobile (≤ 480px): 더 축소 ─── */
@media (max-width: 480px) {
  html, body, .stApp { font-size: 13px !important; }
  .brand-title { font-size: 0.95rem !important; }
  [data-testid="stDataFrame"] { font-size: 0.72rem; }
}
</style>
"""


# ────────────────────────────────────────────────────────────────
# Data loaders (cached)
# ────────────────────────────────────────────────────────────────

@st.cache_data(ttl=30)
def load_counts(db_path: str, since_date: str | None):
    return database.count_by_source(db_path, since_date=since_date)


@st.cache_data(ttl=30)
def load_rows(db_path: str, since_date: str | None, bid_types: tuple[str, ...],
              keyword: str | None, limit: int,
              org_name: str | None = None,
              sources: tuple[str, ...] = ()):
    return database.fetch_for_dashboard(
        db_path,
        since_date=since_date,
        bid_types=list(bid_types) if bid_types else None,
        keyword=keyword or None,
        org_name=org_name or None,
        sources=list(sources) if sources else None,
        limit=limit,
    )


@st.cache_data(ttl=30)
def load_trend(db_path: str, days: int):
    return database.daily_counts(db_path, days=days)


@st.cache_data(ttl=30)
def load_db_meta(db_path: str):
    if not Path(db_path).exists():
        return None
    mtime = datetime.fromtimestamp(Path(db_path).stat().st_mtime)
    return mtime


def invalidate_all_caches():
    load_counts.clear()
    load_rows.clear()
    load_trend.clear()
    load_db_meta.clear()


# ────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────

def _fix_kepco_url(url):
    """KEPCO filenlink는 http:// 로 반환되는 경우가 있음. 클라우드 앱이 HTTPS라
    혼합콘텐츠 경고로 브라우저가 막으므로 강제 HTTPS 변환 (safety net)."""
    if isinstance(url, str) and url.startswith("http://srm.kepco.net"):
        return "https://" + url[7:]
    return url


def _fix_alio_url(url, title) -> str | None:
    """ALIO bidView.do URLs are 404 (legacy). Rewrite to the search-URL form
    at READ time so this works even if the DB migration didn't run.

    pandas가 빈 값을 NaN으로 바꾸므로 str 타입 가드 필수.
    """
    if not isinstance(url, str) or not url:
        return url if isinstance(url, str) else None
    if "alio.go.kr" in url and "bidView.do" in url:
        import urllib.parse
        kw = (title if isinstance(title, str) else "")[:30]
        return (
            "https://www.alio.go.kr/occasional/bidList.do"
            f"?type=title&word={urllib.parse.quote(kw)}"
        )
    return url


def _fix_prvt_url(url, source, bid_no, title) -> str | None:
    """누리장터(prvt_api_*) rows 의 detail_url 정규화.

    다음 경우엔 PDF 공고문 URL 로 재구성:
      1. url 이 비어있음 (None / "")
      2. url 이 레거시 Google 검색 URL 로 저장됨 (초기 커밋 당시 생성된 값)

    PDF URL 패턴 (공개, SSO 불필요):
      https://www.g2b.go.kr/pn/pnp/pnpe/UntyAtchFile/downloadFile.do
        ?bidPbancNo={bidNtceNo}&bidPbancOrd={000-padded}
        &fileType=&fileSeq=1&prcmBsneSeCd=22

    이미 정상인 g2b PDF URL 은 그대로 유지.
    """
    if not isinstance(source, str) or not source.startswith("prvt_api"):
        return url  # not a prvt row — leave as-is

    is_empty = (not isinstance(url, str) or not url.strip())
    is_google = (isinstance(url, str)
                 and "google.com/search" in url)
    if not (is_empty or is_google):
        return url

    if not isinstance(bid_no, str) or not bid_no:
        return None
    if "-" in bid_no:
        core_no, ord_part = bid_no.split("-", 1)
    else:
        core_no, ord_part = bid_no, "000"
    ord_part = (ord_part or "000").zfill(3)[-3:]
    return (
        "https://www.g2b.go.kr/pn/pnp/pnpe/UntyAtchFile/downloadFile.do"
        f"?bidPbancNo={core_no}&bidPbancOrd={ord_part}"
        f"&fileType=&fileSeq=1&prcmBsneSeCd=22"
    )


def _parse_open_date(s):
    """open_date 문자열을 date로 파싱. 소스별 포맷 다양:
      G2B/누리: '2026-04-22 10:00', '2026-04-22 10:00:00'
      ALIO:     '2026.04.23'
      KEPCO:    '2026-04-08 10:00:00' / 'noticeDate' '20260401'
    """
    if not isinstance(s, str) or not s.strip():
        return None
    s = s.strip()
    # Try common formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(s[:len(datetime.now().strftime(fmt))], fmt).date()
        except (ValueError, TypeError):
            continue
    # Fallback: first 10 chars as YYYY-MM-DD or YYYY.MM.DD
    head = s[:10].replace(".", "-").replace("/", "-")
    try:
        return datetime.strptime(head, "%Y-%m-%d").date()
    except ValueError:
        return None


def _is_new_today(open_date_str, today=None):
    """open_date 가 오늘(수집 기준일) 인 공고면 True."""
    today = today or date.today()
    d = _parse_open_date(open_date_str)
    return d is not None and d == today


def _to_eok(price):
    """정수/실수 가격 → 억 단위 float, None/0/NaN → None (정렬 시 최하단으로)."""
    import math
    try:
        p = float(price) if price is not None else None
    except (TypeError, ValueError):
        return None
    if p is None or (isinstance(p, float) and math.isnan(p)) or p <= 0:
        return None
    return round(p / EOK, 2)


def rows_to_dataframe(rows: list[dict]) -> pd.DataFrame:
    cols = ["신규", "bid_no", "title", "org_name", "금액(억원)", "close_date",
            "bid_type", "source_label", "detail_url"]
    if not rows:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(rows)
    if "estimated_price" in df.columns:
        df["금액(억원)"] = df["estimated_price"].apply(_to_eok)
    if "source" in df.columns:
        df["source_label"] = df["source"].map(SOURCE_LABELS).fillna(df["source"])
    if "detail_url" in df.columns and "title" in df.columns:
        sources_col = df["source"] if "source" in df.columns else [""] * len(df)
        bidnos_col  = df["bid_no"] if "bid_no" in df.columns else [""] * len(df)
        df["detail_url"] = [
            _fix_kepco_url(_fix_alio_url(
                _fix_prvt_url(u, s, b, t), t
            ))
            for u, s, b, t in zip(
                df["detail_url"], sources_col, bidnos_col, df["title"]
            )
        ]
    # "신규" 배지 — 오늘 올라온 공고만 "N"
    today = date.today()
    if "open_date" in df.columns:
        df["신규"] = df["open_date"].apply(
            lambda s: "new" if _is_new_today(s, today) else ""
        )
    else:
        df["신규"] = ""
    return df[[c for c in cols if c in df.columns]]


def df_to_excel_bytes(df: pd.DataFrame,
                        sheet_name: str = "입찰공고") -> bytes:
    """Convert displayed dataframe → formatted XLSX bytes.

    헤더 한글화, 칼럼 순서 정리, 금액은 숫자 포맷, 상세보기는 하이퍼링크,
    열 너비 자동 조정.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 한글 헤더 매핑 (dashboard 컬럼명 → 엑셀 표기)
    header_map = {
        "신규": "신규",
        "bid_no": "공고번호",
        "title": "제목",
        "org_name": "공고기관",
        "금액(억원)": "금액(억원)",
        "close_date": "마감일시",
        "bid_type": "업종",
        "source_label": "출처",
        "detail_url": "상세보기(링크)",
    }
    # 컬럼 순서 고정
    ordered = [c for c in header_map if c in df.columns]
    out_df = df[ordered].copy()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 헤더
    thin = Side(border_style="thin", color="D6D3CE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F1EFE8")  # warm soft
    header_font = Font(bold=True, color="3E2C20")         # coffee text
    new_fill    = PatternFill("solid", fgColor="FFF59D")  # NEW 배지 (연노랑)
    new_font    = Font(bold=True, color="000000")

    for col_idx, col in enumerate(ordered, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_map[col])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 데이터 행
    for r_idx, row in enumerate(out_df.itertuples(index=False), start=2):
        for c_idx, (col_name, val) in enumerate(zip(ordered, row), start=1):
            # NaN/None/float-NaN 을 깔끔히 처리
            if val is None:
                v = None
            elif isinstance(val, float):
                import math
                v = None if math.isnan(val) else val
            else:
                v = val
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_name == "금액(억원)":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name == "신규" and v == "new":
                cell.fill = new_fill
                cell.font = new_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "detail_url" and isinstance(v, str) and v.startswith("http"):
                cell.hyperlink = v
                cell.value = "열기"
                cell.font = Font(color="0563C1", underline="single")

    # 열 너비 자동 (간이 heuristic)
    widths = {"신규": 6, "bid_no": 18, "title": 55, "org_name": 28,
              "금액(억원)": 12, "close_date": 20, "bid_type": 8,
              "source_label": 14, "detail_url": 10}
    for c_idx, col in enumerate(ordered, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col, 14)

    ws.freeze_panes = "A2"  # 첫 행 고정
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def humanize_since(dt: datetime) -> str:
    delta = datetime.now() - dt
    if delta.total_seconds() < 60:
        return "방금 전"
    m = int(delta.total_seconds() / 60)
    if m < 60:
        return f"{m}분 전"
    h = m // 60
    if h < 24:
        return f"{h}시간 전"
    return f"{h // 24}일 전"


def render_kw_chips(include: list[str], exclude: list[str]) -> str:
    parts = []
    for kw in include:
        parts.append(f'<span class="kw-chip">+ {kw}</span>')
    for kw in exclude:
        parts.append(f'<span class="kw-chip ex">- {kw}</span>')
    return " ".join(parts) if parts else "<span style='color:#6b7280'>(설정된 키워드 없음)</span>"


# ────────────────────────────────────────────────────────────────
# Actions (in-dashboard collect / notify)
# ────────────────────────────────────────────────────────────────

def run_collect_action(config: dict, db_path: Path,
                        log_callback=None,
                        lookback_override: int | None = None,
                        status_update=None) -> tuple[bool, str, list[str]]:
    """Run collection in parallel (ThreadPoolExecutor).

    lookback_override 가 지정되면 config 의 lookback_days 대신 사용.
    대시보드에서 사용자가 슬라이더로 즉석 지정 가능.
    """
    import time as _time
    from concurrent.futures import ThreadPoolExecutor
    from collectors import (g2b_api, alio_crawler,
                            d2b_api, kwater_api, kepco_api, prvt_api)
    from db import database as dbmod

    sleep = float(config.get("collection", {}).get("request_sleep_seconds", 0.5))
    page_size = int(config.get("collection", {}).get("page_size", 100))
    lookback = int(lookback_override if lookback_override is not None
                   else config.get("collection", {}).get("lookback_days", 1))
    sources = config.get("collection", {}).get("sources") or {}

    dbmod.init_db(db_path)
    total = 0
    errors: list[str] = []
    log_lines: list[str] = []

    def _log(msg: str):
        log_lines.append(msg)
        if log_callback:
            try:
                log_callback(msg)
            except Exception:
                pass

    # 1. 각 활성 소스의 수집 함수를 (이름, 함수) 튜플로 수집
    tasks: list[tuple[str, callable]] = []

    if sources.get("g2b_api"):
        key = get_secret("G2B_SERVICE_KEY")
        if not key:
            msg = "❌ 나라장터(G2B): G2B_SERVICE_KEY 미설정 — Secrets에 추가하세요."
            errors.append(msg); _log(msg)
        else:
            tasks.append(("나라장터(G2B)", lambda k=key: g2b_api.collect_all(
                service_key=k, page_size=page_size,
                sleep_seconds=sleep, lookback_days=lookback,
            )))

    if sources.get("prvt_api"):
        key = get_secret("G2B_SERVICE_KEY")
        if not key:
            _log("⏩ 누리장터(민간): G2B 키 없음 — skip")
        else:
            tasks.append(("누리장터(민간)", lambda k=key: prvt_api.collect_all(
                service_key=k, page_size=page_size,
                sleep_seconds=sleep, lookback_days=lookback,
            )))

    if sources.get("alio"):
        alio_cfg = (config.get("collection", {}).get("alio") or {})
        tasks.append(("ALIO", lambda c=alio_cfg: alio_crawler.collect(
            word=c.get("keyword", ""),
            max_pages=int(c.get("max_pages", 10)),
            sleep_seconds=sleep,
            lookback_days=lookback,
        )))

    if sources.get("d2b_api"):
        key = get_secret("G2B_SERVICE_KEY") or get_secret("D2B_SERVICE_KEY")
        if not key:
            _log("⏩ 방위사업청(d2b): 키 없음 — skip")
        else:
            tasks.append(("방위사업청(d2b)", lambda k=key: d2b_api.collect_all(
                service_key=k, page_size=page_size,
                sleep_seconds=sleep, lookback_days=lookback,
            )))

    if sources.get("kwater_api"):
        key = get_secret("G2B_SERVICE_KEY") or get_secret("KWATER_SERVICE_KEY")
        kw_cfg = (config.get("collection", {}).get("kwater") or {})
        if not key or not kw_cfg.get("base_url"):
            _log("⏩ K-water: 키 또는 base_url 미설정 — skip")
        else:
            tasks.append(("K-water", lambda k=key, c=kw_cfg: kwater_api.collect(
                service_key=k,
                base_url=c["base_url"],
                type_param=c.get("type_param", "_type"),
                page_size=page_size,
                sleep_seconds=sleep,
                lookback_days=lookback,
            )))

    if sources.get("kepco_api"):
        kepco_key = get_secret("KEPCO_API_KEY")
        kepco_cfg = (config.get("collection", {}).get("kepco") or {})
        if not kepco_key:
            _log("⏩ KEPCO: KEPCO_API_KEY 없음 — skip (bigdata.kepco.co.kr 발급 필요)")
        else:
            tasks.append(("KEPCO", lambda k=kepco_key, c=kepco_cfg: kepco_api.collect(
                api_key=k,
                base_url=c.get("base_url") or kepco_api.DEFAULT_BASE_URL,
                company_ids=c.get("company_ids") or None,
                sleep_seconds=sleep,
                lookback_days=lookback,
            )))

    if not tasks:
        summary = "수집할 소스가 없습니다."
        return False, summary, log_lines

    # 2. 병렬 실행 — 각 소스 독립적이라 thread-safe
    #    워커 스레드에서는 Streamlit UI를 직접 못 부르므로 queue에 메시지를 쌓고
    #    메인 스레드가 주기적으로 drain → _log() 호출해 실시간 진행 표시.
    import queue as _queue
    msg_queue: _queue.Queue = _queue.Queue()

    _log(f"⏳ {len(tasks)}개 소스 병렬 수집 시작 — 완료되는 순서대로 표시됩니다")
    for nm, _fn in tasks:
        _log(f"  · {nm} 대기열 등록")
    t0 = _time.time()
    results: list[tuple[str, list, float, Exception | None]] = []

    def _run_one(name, fn):
        msg_queue.put(f"⏳ {name} 수집 중…")
        ts = _time.time()
        try:
            rows = fn()
            elapsed = _time.time() - ts
            msg_queue.put(f"✅ {name}: {len(rows):,}건 수집 ({elapsed:.0f}s)")
            return (name, rows, elapsed, None)
        except Exception as e:
            elapsed = _time.time() - ts
            msg_queue.put(f"❌ {name} 오류 ({elapsed:.0f}s): {type(e).__name__}: {e}")
            return (name, [], elapsed, e)

    def _drain_queue():
        while True:
            try:
                msg = msg_queue.get_nowait()
            except _queue.Empty:
                return
            _log(msg)

    with ThreadPoolExecutor(max_workers=len(tasks)) as pool:
        futures = {pool.submit(_run_one, n, f): n for n, f in tasks}
        pending = set(futures)
        n_total = len(futures)
        while pending:
            _drain_queue()
            done = {f for f in pending if f.done()}
            for f in done:
                try:
                    results.append(f.result())
                except Exception as e:
                    nm = futures.get(f, "?")
                    results.append((nm, [], 0.0, e))
            pending -= done
            # 라이브 진행상황 업데이트 — 매 루프마다 status label 갱신
            if status_update:
                elapsed = _time.time() - t0
                n_done = n_total - len(pending)
                running = [futures[f] for f in pending][:3]
                tail = (" · 진행중: " + ", ".join(running)) if running else ""
                try:
                    status_update(label=(
                        f"⏱ {elapsed:.0f}s 경과 · {n_done}/{n_total} 완료{tail}"
                    ))
                except Exception:
                    pass
            if pending:
                _time.sleep(0.3)
        # 최종 drain (워커 종료 직후 남은 메시지)
        _drain_queue()

    # 3. DB 업서트는 메인 스레드에서 직렬 처리 (SQLite thread-safety 보장)
    _log("💾 DB 저장 중…")
    for name, rows, elapsed, err in results:
        if err:
            errors.append(f"❌ {name} 오류 ({elapsed:.0f}s): {type(err).__name__}: {err}")
        else:
            dbmod.upsert_bids(db_path, rows)
            total += len(rows)

    grand = _time.time() - t0
    summary = f"수집 완료: {total:,}건 (전체 {grand:.0f}s)"
    if errors:
        summary += f" · 오류 {len(errors)}건"
    invalidate_all_caches()
    return len(errors) == 0, summary, log_lines


def run_notify_action(config: dict, db_path: Path, dry_run: bool = True) -> tuple[int, str]:
    from notifiers import email_notifier, slack_notifier
    from db import database as dbmod

    rows = dbmod.get_unnotified(db_path)
    filtered = keyword_filter.apply_filters(rows, config.get("filters") or {})
    if dry_run:
        return len(filtered), f"미리보기: {len(filtered):,}건이 알림 대상입니다. (전송은 하지 않음)"

    channels = (config.get("notifier") or {}).get("channels") or []
    # 수신자: config + data/recipients.json 병합
    notifier_cfg = dict(config.get("notifier") or {})
    email_cfg = dict(notifier_cfg.get("email") or {})
    email_cfg["to_addrs"] = recipients_mod.resolve_to_addrs(config)
    notifier_cfg["email"] = email_cfg

    sent = False
    if "email" in channels:
        sent = email_notifier.send_email(filtered, notifier_cfg) or sent
    if "slack" in channels:
        sent = slack_notifier.send_slack(filtered, notifier_cfg) or sent
    if sent:
        dbmod.mark_notified(db_path, [r["id"] for r in filtered])
        invalidate_all_caches()
    return len(filtered), ("전송 완료" if sent else "전송 실패 (설정을 확인하세요)")


# ────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────

def main() -> None:
    st.set_page_config(
        page_title="국내 입찰공고 현황",
        page_icon="📋",
        layout="wide",
        # 데스크톱: 자동으로 펼쳐짐. 모바일: 자동으로 접힘 + 햄버거 버튼 노출
        initial_sidebar_state="auto",
    )
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    # 모바일 세로 스크롤 갇힘 해결: glide-data-grid(Streamlit dataframe 엔진)가
    # touchmove 에 preventDefault(passive:false) 를 걸어 CSS touch-action 이
    # 우회됨. 두 단계 방어:
    #   1) 캡처 단계에서 grid 핸들러 stopImmediatePropagation
    #   2) window/document/html 모두에 리스너 부착 (어느 것이 먼저 호출되든 잡힘)
    #   3) passive:false 로 등록해 필요 시 preventDefault 직접 호출
    import streamlit.components.v1 as _components
    _components.html("""
    <script>
    (function(){
      try {
        var p = window.parent;
        if (!p || !p.document) return;
        // === Material Icons 폰트 CDN 주입 (parent <head>) ===
        // Streamlit Cloud 모바일에서 자체 폰트가 안 들어와 ligature 가
        // raw 텍스트로 보이는 문제 해결. 부모 document 에 직접 link 추가.
        if (!p.__bidIconsLoaded) {
          p.__bidIconsLoaded = true;
          ['https://fonts.googleapis.com/icon?family=Material+Icons',
           'https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined&display=block'
          ].forEach(function(href){
            var l = p.document.createElement('link');
            l.rel = 'stylesheet'; l.href = href;
            p.document.head.appendChild(l);
          });
        }
        // 매 rerun 마다 재설치 (Streamlit 이 DOM 을 재렌더 할 때 listener 가
        // 남아있어도 중복 호출만 발생 — 상단에 removeEventListener 시도).
        if (p.__bidTouchCleanup) { try { p.__bidTouchCleanup(); } catch(_){} }

        var sx = 0, sy = 0, tracking = false, decided = false, isVertical = false;

        var onStart = function(e){
          if (!e.touches || !e.touches[0]) return;
          sx = e.touches[0].clientX;
          sy = e.touches[0].clientY;
          tracking = true;
          decided = false;
          isVertical = false;
        };
        var onMove = function(e){
          if (!tracking || !e.touches || !e.touches[0]) return;
          var t = e.target;
          var inDf = false;
          while (t) {
            if (t.getAttribute && (
                t.getAttribute('data-testid') === 'stDataFrame' ||
                (t.className && String(t.className).indexOf('dvn-') >= 0) ||
                t.tagName === 'CANVAS')) {
              inDf = true; break;
            }
            t = t.parentNode || (t.host ? t.host : null);
          }
          if (!inDf) return;
          var dx = Math.abs(e.touches[0].clientX - sx);
          var dy = Math.abs(e.touches[0].clientY - sy);
          // 한 번 방향 결정되면 끝까지 유지 (제스처 중간에 흔들림 방지).
          // 결정 임계: 누적 이동 12px 이상 + 명확한 비율 (1.5:1)
          if (!decided) {
            if (dx + dy < 12) return;  // 너무 작은 움직임은 판단 보류
            if (dy > dx * 1.5) {
              isVertical = true;       // 세로 우세 — 페이지 스크롤
            } else if (dx > dy * 1.2) {
              isVertical = false;      // 가로 우세 — grid 컬럼 스크롤
            } else {
              return;                  // 애매하면 한 프레임 더 기다림
            }
            decided = true;
          }
          if (isVertical) {
            // 세로: grid handler 차단 → 브라우저 기본 페이지 스크롤
            e.stopImmediatePropagation();
          }
          // 가로: 아무것도 하지 않음 → grid 가 정상 처리
        };
        var onEnd = function(){ tracking = false; decided = false; };

        // 최대한 상위 (capture) 에서 잡는다
        var targets = [p.window, p.document, p.document.documentElement, p.document.body];
        targets.forEach(function(tg){
          if (!tg) return;
          tg.addEventListener('touchstart', onStart, { capture: true, passive: true });
          tg.addEventListener('touchmove',  onMove,  { capture: true, passive: false });
          tg.addEventListener('touchend',   onEnd,   { capture: true, passive: true });
          tg.addEventListener('touchcancel', onEnd,  { capture: true, passive: true });
        });

        p.__bidTouchCleanup = function(){
          targets.forEach(function(tg){
            if (!tg) return;
            tg.removeEventListener('touchstart', onStart, { capture: true });
            tg.removeEventListener('touchmove',  onMove,  { capture: true });
            tg.removeEventListener('touchend',   onEnd,   { capture: true });
            tg.removeEventListener('touchcancel', onEnd,  { capture: true });
          });
        };
      } catch (err) { /* ignore */ }
    })();
    </script>
    """, height=0)

    config = load_config(ROOT / "config.yaml")
    db_path = ROOT / (config.get("database", {}).get("path") or "data/bids.sqlite")

    # Always init DB (idempotent — CREATE TABLE IF NOT EXISTS) +
    # run migrations. 빈 DB 라도 사이드바·UI 가 정상 렌더되도록.
    try:
        database.init_db(db_path)
        invalidate_all_caches()
    except Exception:
        pass

    # ── Secret availability check (warn once per session) ──
    if not get_secret("G2B_SERVICE_KEY"):
        st.warning(
            "⚠️ **나라장터(G2B) API 키가 설정되지 않았습니다.** "
            "Streamlit Cloud 앱 설정 > Secrets 탭에 `G2B_SERVICE_KEY = \"...\"` 추가하세요. "
            "현재는 ALIO 데이터만 수집됩니다."
        )

    # ── Brand bar ──────────────────────────────────────────
    db_mtime = load_db_meta(str(db_path))
    last_update = humanize_since(db_mtime) if db_mtime else "없음"
    st.markdown(
        f"""
        <div class="brand-bar">
          <div class="brand-title">국내 입찰공고 현황</div>
          <div class="brand-sub">마지막 수집 · <b>{last_update}</b></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── First-run notice (DB 비어있어도 사이드바·UI 는 정상 렌더) ──
    # st.stop() 을 쓰면 사이드바가 등록되지 않아 모바일 햄버거 버튼이
    # 사라짐. 대신 안내만 표시하고 계속 진행 (사이드바의 '지금 수집' 사용).
    if load_counts(str(db_path), None) == {}:
        st.info(
            "아직 수집된 데이터가 없습니다. "
            "👈 좌측 사이드바의 **지금 수집** 버튼을 눌러주세요. "
            "(모바일은 좌상단 햄버거 → 사이드바)"
        )

    # ── Session state defaults ──
    # Live-reactive: 모든 필터 입력은 f_*_input key로 session_state에 자동 바인드.
    today = date.today()
    cfg_filters = config.get("filters", {})
    cfg_include = cfg_filters.get("include_keywords", [])
    cfg_exclude = cfg_filters.get("exclude_keywords", [])
    bid_type_options = ["물품", "용역", "공사", "외자", "기타", "민간"]
    cfg_types = cfg_filters.get("bid_types") or bid_type_options[:3]
    for t in cfg_types:
        if t not in bid_type_options:
            bid_type_options.append(t)
    default_types = [t for t in cfg_types if t in bid_type_options]

    _init_min_eok = int(cfg_filters.get("min_amount_eok") or 0)
    _init_max_eok = int(cfg_filters.get("max_amount_eok") or 9999)
    _filter_defaults = {
        "f_bid_types_input": default_types,
        "f_keyword_input": "",
        "f_org_query_input": "",
        "f_include_text_input": ", ".join(cfg_include),
        "f_exclude_text_input": ", ".join(cfg_exclude),
        "f_amount_slider_input": (_init_min_eok, _init_max_eok),
        # 공고일 범위 — 기본: 최근 30일
        "f_date_range_input": (today - timedelta(days=30), today),
    }
    _misc_defaults = {
        # 그룹별 체크박스 (기본: 전부 체크 → 전체 표시)
        "mcard_나라장터": True,
        "mcard_누리장터": True,
        "mcard_기타": True,
        # 수집 범위 — 기본 7일
        "f_collect_lookback_input": int(config.get("collection", {}).get("lookback_days", 30)),
        # 마감 안 지난 활성 공고만 표시 (기본 ON)
        "f_active_only_input": True,
    }
    for _k, _v in {**_filter_defaults, **_misc_defaults}.items():
        st.session_state.setdefault(_k, _v)

    # ── 인허가 탭은 임시 숨김 (아래 블록을 `if True:` 로 바꾸면 복원) ──
    # tab_bid, tab_permit = st.tabs(["입찰 공고", "인허가 현황"])
    # _tab_ctx = tab_bid.__enter__()

    # ── Section 1: 상단 그룹 필터 체크박스 (중복 선택 가능) ──
    today_counts = load_counts(str(db_path), today.isoformat())
    total_counts = load_counts(str(db_path), None)

    group_counts = {g: sum(today_counts.get(s, 0) for s in srcs)
                    for g, srcs in SOURCE_GROUPS.items()}
    total_today = sum(today_counts.values())

    # Streamlit은 이미 렌더된 widget key에 직접 할당 불가 → on_click 콜백 사용.
    def _select_all_groups():
        for _g in SOURCE_GROUPS:
            st.session_state[f"mcard_{_g}"] = True

    def _clear_all_groups():
        for _g in SOURCE_GROUPS:
            st.session_state[f"mcard_{_g}"] = False

    # 전체 섹션을 container 로 감싸 모바일 컴팩트 CSS 스코프 만듦
    with st.container(key="src_filter_section"):
        st.markdown("<div class='src-section-title'>오늘의 수집 현황</div>",
                    unsafe_allow_html=True)
        cbx_cols = st.columns([1.3, 1.3, 1.3, 1, 1])
        for i, g in enumerate(SOURCE_GROUPS.keys()):
            cbx_cols[i].checkbox(f"{g} ({group_counts.get(g, 0):,})",
                                  key=f"mcard_{g}")
        cbx_cols[3].button("전체", width="stretch", key="mcard_all_btn",
                            on_click=_select_all_groups,
                            help="3개 그룹 모두 체크 (전체 표시와 동일)")
        cbx_cols[4].button("해제", width="stretch", key="mcard_clear_btn",
                            on_click=_clear_all_groups,
                            help="모두 해제 (= 전체 표시)")

    # 체크된 그룹들의 소스 유니언 → source_filter
    # - 전부 체크 (= 전체 버튼) : 전체 표시
    # - 아무것도 체크 안됨     : 0건 (__NONE__ 이라는 존재하지 않는 소스로 필터)
    # - 부분 체크              : 해당 그룹만
    _checked = [g for g in SOURCE_GROUPS
                if st.session_state.get(f"mcard_{g}", False)]
    if set(_checked) == set(SOURCE_GROUPS.keys()):
        st.session_state["source_filter"] = []
        current_sources = []
    elif not _checked:
        # 존재하지 않는 source 값을 넣어 DB WHERE 절에서 0건 보장
        st.session_state["source_filter"] = ["__NONE__"]
        current_sources = ["__NONE__"]
    else:
        srcs = []
        for g in _checked:
            srcs.extend(SOURCE_GROUPS[g])
        st.session_state["source_filter"] = srcs
        current_sources = srcs

    # 요약 캡션
    group_totals = {g: sum(total_counts.get(s, 0) for s in srcs)
                    for g, srcs in SOURCE_GROUPS.items()}
    st.markdown(
        f"<div class='section-hint'>오늘 <b>{total_today:,}건</b> · "
        f"DB 전체 <b>{sum(total_counts.values()):,}건</b> — "
        + " · ".join(f"{g} {v:,}" for g, v in group_totals.items())
        + "</div>",
        unsafe_allow_html=True,
    )

    # ── Sidebar (live-reactive 필터) ────────────────────────
    with st.sidebar:
        st.markdown("### 필터")

        # 필터 위젯: 모두 key=로 session_state 자동 바인드 → 입력 변경 시 즉시 반영
        st.multiselect("업종", bid_type_options, key="f_bid_types_input")
        st.text_input("공고명 검색 (제목에만 적용)",
                      key="f_keyword_input", placeholder="예: 전기공사")
        st.text_input("기관명 검색",
                      key="f_org_query_input", placeholder="예: 교육청")

        st.markdown("**공고명 포함 키워드** (제목 기준, 하나라도 있으면 통과)")
        st.text_area("include_keywords", key="f_include_text_input", height=70,
                     help="쉼표로 구분, 공고 제목에만 적용",
                     label_visibility="collapsed")
        st.markdown("**공고명 제외 키워드** (제목 기준, 하나라도 있으면 제외)")
        st.text_area("exclude_keywords", key="f_exclude_text_input", height=60,
                     label_visibility="collapsed")

        preview_inc = [k.strip() for k in st.session_state["f_include_text_input"].split(",") if k.strip()]
        preview_exc = [k.strip() for k in st.session_state["f_exclude_text_input"].split(",") if k.strip()]
        st.markdown(render_kw_chips(preview_inc, preview_exc),
                    unsafe_allow_html=True)

        st.slider("금액 범위 (억원)", 0, 9999, key="f_amount_slider_input", step=1)

        # 마감 안 지난 입찰 진행중 공고만 표시 (기본 ON)
        st.checkbox("입찰 진행중인 공고만",
                    key="f_active_only_input",
                    help="마감일이 오늘 이후인 공고만 표시 (마감 지난 공고 제외)")

        # 필터 기본값 복원 — on_click 콜백 (widget key에 직접 할당 불가)
        def _reset_filters_cb():
            for _k, _v in _filter_defaults.items():
                st.session_state[_k] = _v
            invalidate_all_caches()

        st.button("기본값 복원", width="stretch", key="reset_filters",
                  on_click=_reset_filters_cb,
                  help="모든 필터 입력을 config.yaml의 기본값으로 복원 (체크박스는 유지)")

        st.markdown("---")
        st.markdown("### 작업")

        # 공고일 범위 (조회하기 바로 위)
        st.date_input("공고일 범위", key="f_date_range_input",
                      help="선택한 공고 발행일 범위 안의 공고만 표시")

        # 조회하기 — 캐시 비우고 재조회 (현재 필터 즉시 재평가)
        if st.button("조회하기", width="stretch", type="primary",
                     key="refetch_btn"):
            invalidate_all_caches()
            st.rerun()

        # 수집 범위 — 슬라이더로 사용자 지정 (기본 30일, 누리장터 등 마감
        # 안 지난 활성 공고 충분히 잡으려면 30일 권장)
        st.slider("수집 범위 (일)", min_value=1, max_value=90, value=30,
                  key="f_collect_lookback_input",
                  help="오늘 기준 과거 N일 동안 올라온 공고를 수집")

        # 수집하기 (아래) — 외부 API 호출
        if st.button("지금 수집", width="stretch", type="secondary",
                     key="collect_btn"):
            lookback_v = int(st.session_state.get("f_collect_lookback_input", 30))
            with st.status(f"공고 수집 중 (최근 {lookback_v}일)… 약 1~2분 소요됩니다.",
                           expanded=True) as status:
                def _stream(msg: str):
                    st.write(msg)
                ok, summary, _logs = run_collect_action(
                    config, db_path, log_callback=_stream,
                    lookback_override=lookback_v,
                    status_update=status.update,  # 라이브 타이머
                )
                status.update(label=summary,
                              state=("complete" if ok else "error"))
            if not ok:
                st.error("일부 소스 수집 실패. 위 로그를 확인하세요.")
            st.rerun()

        if st.button("알림 미리보기 (dry-run)", width="stretch"):
            count, msg = run_notify_action(config, db_path, dry_run=True)
            st.success(msg) if count else st.info(msg)

        # ── 메일링 수신자 관리 ──
        st.markdown("---")
        st.markdown("### 메일링")
        with st.expander("수신자 관리", expanded=False):
            current_recipients = recipients_mod.load()
            cfg_recipients = ((config.get("notifier") or {}).get("email", {})
                              .get("to_addrs") or [])
            all_recipients = recipients_mod.resolve_to_addrs(config)

            if all_recipients:
                st.caption(f"현재 수신자 {len(all_recipients)}명")
                for email in all_recipients:
                    is_removable = email in current_recipients
                    c1, c2 = st.columns([5, 1])
                    with c1:
                        tag = "" if is_removable else " · config"
                        st.markdown(f"<div style='font-size:0.85rem; padding:4px 0'>"
                                    f"{email}<span style='color:var(--fg-muted)'>{tag}</span>"
                                    f"</div>", unsafe_allow_html=True)
                    with c2:
                        if is_removable:
                            if st.button("삭제", key=f"rm_{email}",
                                          help="이 수신자를 목록에서 제거"):
                                recipients_mod.remove(email)
                                st.rerun()
            else:
                st.caption("등록된 수신자가 없습니다.")

            def _add_recipient_cb():
                email = (st.session_state.get("new_recipient_input") or "").strip()
                if not email:
                    st.session_state["recipient_msg"] = ("error", "이메일을 입력하세요.")
                    return
                if not recipients_mod.is_valid_email(email):
                    st.session_state["recipient_msg"] = ("error",
                        f"올바른 이메일 형식이 아닙니다: {email}")
                    return
                if recipients_mod.add(email):
                    st.session_state["recipient_msg"] = ("success", f"추가됨: {email}")
                    st.session_state["new_recipient_input"] = ""
                else:
                    st.session_state["recipient_msg"] = ("info", f"이미 등록되어 있습니다: {email}")

            st.text_input("이메일 추가",
                          key="new_recipient_input",
                          placeholder="name@example.com")
            st.button("추가", on_click=_add_recipient_cb,
                      width="stretch", key="add_recipient_btn")

            _msg = st.session_state.pop("recipient_msg", None)
            if _msg:
                level, text = _msg
                {"error": st.error, "success": st.success,
                 "info": st.info}.get(level, st.info)(text)

            if st.button("지금 테스트 발송", width="stretch",
                          key="test_send_btn", type="primary",
                          disabled=not all_recipients):
                from notifiers import email_notifier
                sample_row = [{
                    "bid_no": "TEST-001",
                    "title": "메일링 테스트 공고",
                    "org_name": "테스트 기관",
                    "estimated_price": 100_000_000,
                    "close_date": "2026-12-31 23:59",
                    "bid_type": "테스트",
                    "detail_url": None,
                }]
                # Override to_addrs with merged list
                notifier_cfg = dict(config.get("notifier") or {})
                email_cfg = dict(notifier_cfg.get("email") or {})
                email_cfg["to_addrs"] = all_recipients
                notifier_cfg["email"] = email_cfg
                ok = email_notifier.send_email(sample_row, notifier_cfg)
                if ok:
                    st.success(f"테스트 메일 {len(all_recipients)}명에게 발송 완료")
                else:
                    st.error("발송 실패 — SMTP 설정(SMTP_USER/PASS) 확인 필요")

        st.caption(f"DB: `{db_path.name}` · 최종 업데이트 {last_update}")

    # ── Section 2: 필터 적용된 목록 (live-reactive) ──
    # 표 위에 '윈도우 타이틀바' 스타일 컨테이너: 좌측 카운트 + 우측 xlsx 아이콘

    # DB 전체 조회 (공고일 필터는 Python-side에서 적용, 소스별 날짜 포맷이 달라서)
    since_str = None
    bid_types_v = st.session_state["f_bid_types_input"]
    keyword_v = st.session_state["f_keyword_input"]
    org_v = st.session_state["f_org_query_input"]
    applied_include = [k.strip() for k in st.session_state["f_include_text_input"].split(",") if k.strip()]
    applied_exclude = [k.strip() for k in st.session_state["f_exclude_text_input"].split(",") if k.strip()]
    lo_hi = st.session_state["f_amount_slider_input"]
    min_eok, max_eok = lo_hi[0], lo_hi[1]
    row_limit = 1_000_000  # 제한 없음 (사용자 요청)

    # 공고일 범위 파싱 (st.date_input range는 tuple/list 또는 single date)
    _date_val = st.session_state.get("f_date_range_input")
    d_start = d_end = None
    if isinstance(_date_val, (tuple, list)):
        if len(_date_val) >= 1: d_start = _date_val[0]
        if len(_date_val) >= 2: d_end = _date_val[1]
    elif isinstance(_date_val, date):
        d_start = _date_val

    rows = load_rows(
        str(db_path), since_str, tuple(bid_types_v), keyword_v,
        row_limit, org_name=org_v,
        sources=tuple(current_sources) if current_sources else (),
    )

    # 공고일 범위 필터 (Python-side) — open_date 포맷이 소스마다 달라서
    if d_start or d_end:
        def _in_date_range(r: dict) -> bool:
            d = _parse_open_date(r.get("open_date"))
            if d is None:
                return True  # 파싱 실패한 공고는 보수적으로 포함
            if d_start and d < d_start: return False
            if d_end and d > d_end: return False
            return True
        rows = [r for r in rows if _in_date_range(r)]

    # 입찰 진행중 (마감일 ≥ 오늘) 필터
    if st.session_state.get("f_active_only_input", True):
        _today = date.today()
        def _is_active(r: dict) -> bool:
            d = _parse_open_date(r.get("close_date"))
            if d is None:
                return True   # 마감일 파싱 실패 시 보수적으로 포함
            return d >= _today
        rows = [r for r in rows if _is_active(r)]

    filter_cfg = {
        "include_keywords": applied_include,
        "exclude_keywords": applied_exclude,
        "min_amount_eok": min_eok,
        "max_amount_eok": max_eok,
        "case_sensitive": bool(cfg_filters.get("case_sensitive", False)),
    }
    rows = keyword_filter.apply_filters(rows, filter_cfg)

    df = rows_to_dataframe(rows)

    # 타이틀바 — horizontal 컨테이너 (Streamlit 1.38+)
    # 좌측: 📋 공고 리스트 + 검색 결과 카운트 / 우측: ⤓ xlsx
    with st.container(
        horizontal=True,
        horizontal_alignment="distribute",
        vertical_alignment="center",
        gap=None,
        key="bidtable_titlebar",
    ):
        st.markdown(
            f"<div class='tbl-title'>📋 <b>입찰 공고 리스트</b> · "
            f"검색 결과 <b>{len(df):,}</b>건</div>",
            unsafe_allow_html=True,
        )
        if len(df) > 0:
            try:
                xlsx_bytes = df_to_excel_bytes(df)
                st.download_button(
                    "⤓ xlsx",
                    data=xlsx_bytes,
                    file_name=f"bids_{date.today().isoformat()}.xlsx",
                    mime=("application/vnd.openxmlformats-officedocument"
                          ".spreadsheetml.sheet"),
                    key="excel_dl_btn",
                    help="현재 필터가 적용된 검색 결과를 엑셀로 저장",
                )
            except Exception as e:
                st.caption(f"xlsx 변환 실패: {type(e).__name__}")

    if len(df) > 0:
        # NEW 배지: "N" 값에 연노랑 배경 + 검정 볼드 + 가운데 정렬.
        # pandas Styler 는 st.dataframe 에서 CSS 로 렌더됨.
        # pandas 2.1+ 는 Styler.map(), 이전 버전은 Styler.applymap() — 버전 호환.
        def _style_new(v):
            if v == "new":
                return ("background-color: #FFF59D; color: #000000; "
                        "font-weight: 800; text-align: center;")
            return ""
        _styler = df.style
        _apply = getattr(_styler, "map", None) or _styler.applymap
        styled = _apply(_style_new, subset=["신규"])

        st.dataframe(
            styled,
            width="stretch",
            # 행이 많을수록 표 자체도 길게 (글이드 데이터그리드는 가상 스크롤이라 OK).
            # 모바일은 CSS @media 로 calc(100vh - X) 강제하여 viewport 채움.
            height=min(900, max(450, len(df) * 36 + 60)),
            hide_index=True,
            column_config={
                "신규": st.column_config.TextColumn(
                    "신규", width=55,  # 픽셀 단위 — small(~100)의 약 절반
                    help="오늘 올라온 공고는 'new' 로 표시",
                ),
                "bid_no": st.column_config.TextColumn("공고번호", width="small"),
                "title": st.column_config.TextColumn("제목", width="large"),
                "org_name": st.column_config.TextColumn("기관", width="medium"),
                "금액(억원)": st.column_config.NumberColumn(
                    "금액(억원)", format="%.2f", width="small",
                    help="빈 칸은 공고문 내 '금액 링크 참조'. 정렬 시 자동으로 최하단.",
                ),
                "close_date": st.column_config.TextColumn("마감", width="small"),
                "bid_type": st.column_config.TextColumn("업종", width="small"),
                "source_label": st.column_config.TextColumn("출처", width="small"),
                "detail_url": st.column_config.LinkColumn(
                    "상세보기", display_text="🔗 열기", width="small"
                ),
            },
        )
    else:
        hint = []
        if applied_include:
            hint.append(f"포함 키워드({', '.join(applied_include)})와 겹치는 제목이 없을 수 있습니다")
        if bid_types_v:
            hint.append("업종 선택을 바꿔보세요")
        if keyword_v:
            hint.append(f"공고명 검색 '{keyword_v}' 제외해보세요")
        if org_v:
            hint.append(f"기관명 검색 '{org_v}' 제외해보세요")
        if current_sources == ["__NONE__"]:
            hint.append("상단 체크박스를 선택하거나 '전체' 버튼을 누르세요")
        elif current_sources:
            hint.append("상단 다른 그룹도 체크하거나 '전체' 버튼을 눌러보세요")
        st.markdown(
            f"""<div class='empty-state'>
               조건에 해당하는 공고가 없습니다.<br>
               <small>{' · '.join(hint) if hint else '필터를 완화해보세요'}</small>
               </div>""",
            unsafe_allow_html=True,
        )

    # ── 입찰 공고 탭 종료, 인허가 현황 탭 시작 ──
    # tab_bid.__exit__(None, None, None)

    if False:  # 인허가 탭 임시 숨김 — 기능 재활성화 시 True 로 변경
        from collectors import permit_api
        from dashboard.permit_regions import REGION_PRESETS

        st.markdown("### 인허가 현황")
        st.markdown(
            "<div class='section-hint'>"
            "국토부 <b>건축HUB 건축인허가정보</b>(ArchPmsHubService)에서 선택한 지역의 "
            "허가·착공·사용승인 실제 데이터를 조회합니다. "
            "<br><small>※ 시군구·법정동 단위 조회 (전국 일괄 조회는 지원 안 됨)</small>"
            "</div>",
            unsafe_allow_html=True,
        )

        permit_key = get_secret("PERMIT_API_KEY") or get_secret("BUILDING_PERMIT_KEY")
        if not permit_key:
            st.warning(
                "PERMIT_API_KEY 가 설정되지 않았습니다. "
                "Streamlit Cloud Secrets 또는 `.env`에 다음을 추가하세요:\n\n"
                "`PERMIT_API_KEY = \"여기에 data.go.kr 건축인허가 일반 인증키\"`"
            )
            st.caption(
                "발급: https://www.data.go.kr/data/15134735/openapi.do → 활용신청 → 승인"
            )
            return  # 인허가 탭만 여기서 종료 (전체 페이지 유지)

        # ── 조회 입력 ──
        preset_labels = ["(직접 입력)"] + [p[0] for p in REGION_PRESETS]
        sel_idx = st.selectbox(
            "지역 선택",
            options=list(range(len(preset_labels))),
            format_func=lambda i: preset_labels[i],
            index=1,  # 기본: 서울 강남구 역삼동
            key="permit_region_idx",
            help="자주 조회하는 지역은 프리셋에서 선택, 그 외는 직접 입력",
        )
        use_custom = (sel_idx == 0)

        if not use_custom:
            # 프리셋 선택 → 코드 자동 결정 (직접 입력 안 받음)
            sigungu_cd, bjdong_cd = REGION_PRESETS[sel_idx - 1][1:]
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                st.text_input("시군구 코드 (sigunguCd)",
                              value=sigungu_cd, disabled=True)
            with col_c2:
                st.text_input("법정동 코드 (bjdongCd)",
                              value=bjdong_cd, disabled=True)
        else:
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                sigungu_cd = st.text_input(
                    "시군구 코드 (sigunguCd)", value="11680",
                    key="permit_sigungu_cd",
                    help="5자리 법정동 상위 코드 (예: 서울 강남구 = 11680)",
                )
            with col_c2:
                bjdong_cd = st.text_input(
                    "법정동 코드 (bjdongCd)", value="10100",
                    key="permit_bjdong_cd",
                    help="5자리 법정동 하위 코드 (예: 역삼동 = 10100)",
                )

        col_d1, col_d2, col_d3 = st.columns([2, 2, 2])
        with col_d1:
            p_start = st.date_input(
                "허가일 시작 (선택)",
                value=None,
                key="permit_start_date",
                help="비우면 제한 없음",
            )
        with col_d2:
            p_end = st.date_input(
                "허가일 종료 (선택)",
                value=None,
                key="permit_end_date",
            )
        with col_d3:
            max_pages = st.number_input(
                "최대 페이지 (100건/page)",
                min_value=1, max_value=50, value=5,
                key="permit_max_pages",
                help="건수가 많은 지역은 늘리세요",
            )

        go = st.button("인허가 조회", type="primary",
                       width="stretch", key="permit_fetch_btn")

        if go:
            if not sigungu_cd or not bjdong_cd:
                st.error("시군구·법정동 코드를 모두 입력하세요.")
            else:
                def _to_yyyymmdd(d):
                    return d.strftime("%Y%m%d") if d else None
                with st.spinner(f"국토부 건축HUB 조회 중… ({sigungu_cd}-{bjdong_cd})"):
                    try:
                        rows = permit_api.collect(
                            service_key=permit_key,
                            sigungu_cd=sigungu_cd.strip(),
                            bjdong_cd=bjdong_cd.strip(),
                            start_date=_to_yyyymmdd(p_start),
                            end_date=_to_yyyymmdd(p_end),
                            page_size=100,
                            max_pages=int(max_pages),
                            sleep_seconds=0.5,
                        )
                    except Exception as e:
                        rows = []
                        st.error(f"조회 실패: {type(e).__name__}: {e}")
                st.session_state["_permit_rows"] = rows
                st.session_state["_permit_query"] = {
                    "sigungu": sigungu_cd, "bjdong": bjdong_cd,
                    "label": (preset_labels[sel_idx]
                              if not use_custom else f"{sigungu_cd}-{bjdong_cd}"),
                }

        rows = st.session_state.get("_permit_rows")
        if rows is not None:
            q = st.session_state.get("_permit_query") or {}
            st.write(f"**{q.get('label','')} · 조회 결과: {len(rows):,}건**")
            if not rows:
                st.markdown(
                    "<div class='empty-state'>"
                    "조회 결과가 없습니다.<br>"
                    "<small>날짜 범위를 넓히거나 다른 지역을 시도해보세요.</small>"
                    "</div>",
                    unsafe_allow_html=True,
                )
            else:
                import pandas as pd
                df_p = pd.DataFrame(rows)
                # 표시용 컬럼 재배열 & 한글 헤더
                display_cols = [
                    "pms_day", "bld_nm", "plat_plc", "main_purps",
                    "strct", "tot_area", "grnd_flr_cnt", "ugrnd_flr_cnt",
                    "hhld_cnt", "stcns_day", "use_apr_day", "mgm_pk",
                ]
                df_p = df_p[[c for c in display_cols if c in df_p.columns]]
                st.dataframe(
                    df_p,
                    width="stretch",
                    hide_index=True,
                    column_config={
                        "pms_day": st.column_config.TextColumn("허가일", width="small"),
                        "bld_nm": st.column_config.TextColumn("건물명", width="medium"),
                        "plat_plc": st.column_config.TextColumn("대지위치", width="large"),
                        "main_purps": st.column_config.TextColumn("주용도", width="small"),
                        "strct": st.column_config.TextColumn("구조", width="small"),
                        "tot_area": st.column_config.NumberColumn(
                            "연면적(㎡)", format="%.1f", width="small",
                        ),
                        "grnd_flr_cnt": st.column_config.NumberColumn(
                            "지상층", format="%d", width="small",
                        ),
                        "ugrnd_flr_cnt": st.column_config.NumberColumn(
                            "지하층", format="%d", width="small",
                        ),
                        "hhld_cnt": st.column_config.NumberColumn(
                            "세대수", format="%d", width="small",
                        ),
                        "stcns_day": st.column_config.TextColumn("착공일", width="small"),
                        "use_apr_day": st.column_config.TextColumn("사용승인", width="small"),
                        "mgm_pk": st.column_config.TextColumn("관리PK", width="small"),
                    },
                )
                # 간단 요약
                with st.expander("요약 통계"):
                    def _sum(k): return sum(r.get(k) or 0 for r in rows)
                    def _cnt(k): return sum(1 for r in rows if r.get(k))
                    st.write({
                        "총 연면적(㎡)": round(_sum("tot_area"), 1),
                        "허가 건수": _cnt("pms_day"),
                        "착공 건수": _cnt("stcns_day"),
                        "사용승인 건수": _cnt("use_apr_day"),
                    })

        with st.expander("API 정보"):
            st.markdown("""
**사용 중인 API**: 국토부 건축HUB 건축인허가정보 서비스
- 엔드포인트: `apis.data.go.kr/1613000/ArchPmsHubService/getApBasisOulnInfo`
- 데이터셋: data.go.kr **15134735**
- 필수 파라미터: `sigunguCd` (시군구), `bjdongCd` (법정동) — 전국 일괄 조회 불가
- 제공 정보: 대지위치, 건물명, 주용도, 구조, 연면적, 층수, 세대수, 허가/착공/사용승인일

**시군구/법정동 코드 조회**: https://www.code.go.kr/stdcode/regCodeL.do
""")


if __name__ == "__main__":
    main()

